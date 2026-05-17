from __future__ import annotations

import openpyxl
from openpyxl.styles import Font, PatternFill

from translator.protein_translation import ORF

MAX_EXCEL_CELL_TEXT = 32767


def _excel_safe_text(value: str) -> str:
    if len(value) <= MAX_EXCEL_CELL_TEXT:
        return value
    suffix = "... [truncated for Excel cell limit]"
    return value[: MAX_EXCEL_CELL_TEXT - len(suffix)] + suffix


def _write_header(ws, headers: list[str]) -> None:
    fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for column, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True)
        cell.fill = fill


def _auto_width(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 80)


def _codon_text(orf: ORF) -> str:
    return " ".join(f"{codon}:{aminoacid}" for codon, aminoacid in orf.codon_pairs)


def export_xlsx(results_dict: dict, filepath: str) -> bool:
    try:
        wb = openpyxl.Workbook()

        summary = wb.active
        summary.title = "Summary"
        _write_header(
            summary,
            [
                "File",
                "Sequence length",
                "GC%",
                "A",
                "T",
                "C",
                "G",
                "Start codons",
                "Stop codons",
                "Longest protein",
            ],
        )

        for row, (filename, data) in enumerate(results_dict.items(), 2):
            stats = data["Stats"]
            counts = stats.base_counts
            values = [
                filename,
                stats.length,
                stats.gc_percent,
                counts["A"],
                counts["T"],
                counts["C"],
                counts["G"],
                stats.start_codon_count,
                stats.stop_codon_count,
                stats.longest_protein_length,
            ]
            for column, value in enumerate(values, 1):
                summary.cell(row=row, column=column, value=value)

        sequences = wb.create_sheet("Sequences")
        _write_header(sequences, ["File", "RNA", "Complement DNA", "Reverse complement DNA"])
        for row, (filename, data) in enumerate(results_dict.items(), 2):
            sequences.cell(row=row, column=1, value=filename)
            sequences.cell(row=row, column=2, value=_excel_safe_text(data.get("RNA", "")))
            sequences.cell(row=row, column=3, value=_excel_safe_text(data.get("Complement", "")))
            sequences.cell(row=row, column=4, value=_excel_safe_text(data.get("ReverseComplement", "")))

        orfs_sheet = wb.create_sheet("ORFs")
        _write_header(
            orfs_sheet,
            [
                "File",
                "Direction",
                "Frame",
                "Start position",
                "Stop position",
                "ORF length",
                "Protein length",
                "Protein",
                "Stop codon",
                "Complete",
                "GC%",
                "Codon:Aminoacid",
            ],
        )

        row = 2
        for filename, data in results_dict.items():
            stats = data["Stats"]
            for orf in data["ORFs"]:
                values = [
                    filename,
                    orf.direction,
                    orf.frame,
                    orf.start_position,
                    orf.stop_position or "",
                    orf.length_nt,
                    orf.protein_length,
                    orf.protein,
                    orf.stop_codon or "",
                    "yes" if orf.complete else "no",
                    stats.gc_percent,
                    _excel_safe_text(_codon_text(orf)),
                ]
                for column, value in enumerate(values, 1):
                    orfs_sheet.cell(row=row, column=column, value=value)
                row += 1

        for ws in wb.worksheets:
            _auto_width(ws)
            ws.freeze_panes = "A2"

        wb.save(filepath)
        return True
    except Exception as exc:
        print(f"[export_xlsx] Failed: {exc}")
        return False
