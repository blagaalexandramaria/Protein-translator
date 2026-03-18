from codon_table import CODON_TABLE
import os
import csv
import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox
import openpyxl
from functools import lru_cache
from concurrent.futures import ThreadPoolExecutor

# –––––––– DNA → RNA & PROTEIN ––––––––
def dna_to_rna(dna: str) -> str:
    return dna.upper().replace("T", "U")

@lru_cache(maxsize=128)
def _lookup_codon(codon: str) -> str:
    return CODON_TABLE.get(codon, "?")

_STOP_CODONS = frozenset({"UAA", "UAG", "UGA"})

def rna_to_protein_all(rna: str) -> list[tuple[str, list, str | None]]:
    proteins = []
    rna_len = len(rna)
    pos = 0
    while pos <= rna_len - 3:
        pos = rna.find("AUG", pos)
        if pos == -1:
            break
        pairs = []
        stop_codon = None
        for i in range(pos, rna_len - 2, 3):
            codon = rna[i:i + 3]
            if len(codon) < 3:
                break
            aa = _lookup_codon(codon)
            if aa == "STOP":
                stop_codon = codon
                break
            pairs.append((codon, aa))
        if pairs:
            protein = "-".join(aa for _, aa in pairs)
            proteins.append((protein, pairs, stop_codon))
        pos += 3
    return proteins

def translate_dna(dna: str) -> tuple[str, list]:
    rna = dna_to_rna(dna)
    proteins = rna_to_protein_all(rna)
    return rna, proteins

# –––––––– CLEAN & VALIDATE ––––––––
_VALID_BASES = frozenset("ATCG")

def clean_dna(dna_input: str) -> str:
    return "".join(c for c in dna_input.upper() if c in _VALID_BASES)

def validate_dna(dna: str) -> bool:
    return bool(dna) and _VALID_BASES.issuperset(dna)

# –––––––– FILE READER ––––––––
def read_file(filepath: str) -> str:
    ext = os.path.splitext(filepath)[1].lower()
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            if ext in (".txt", ".fasta", ".fa"):
                # FASTA: ignorăm liniile ce încep cu ">"
                lines = [line.strip() for line in f if not line.startswith(">")]
                return "".join(lines)
            elif ext == ".csv":
                reader = csv.reader(f)
                return "\n".join(row[0] for row in reader if row)
    except OSError as e:
        print(f"[read_file] Could not read '{filepath}': {e}")
    return ""

# –––––––– PROCESS FILE ––––––––
def process_file(filepath: str, translation_type: str) -> dict:
    content = read_file(filepath)
    if not content:
        return {"RNA": "", "Proteins": []}
    dna = clean_dna(content)
    if not validate_dna(dna):
        return {"RNA": "", "Proteins": []}
    rna, proteins = translate_dna(dna)
    return {
        "RNA": rna if translation_type in ("both", "rna") else "",
        "Proteins": proteins if translation_type in ("both", "protein") else [],
    }

# –––––––– EXPORT XLSX ––––––––
def export_xlsx(results_dict: dict, filepath: str) -> bool:
    try:
        wb = openpyxl.Workbook()
        first_sheet = True

        for filename, data in results_dict.items():
            ws = wb.active if first_sheet else wb.create_sheet()
            ws.title = filename[:31]
            first_sheet = False
            row = 1

            if data["RNA"]:
                ws.cell(row=row, column=1, value="RNA")
                ws.cell(row=row, column=2, value=data["RNA"])
                row += 2

            proteins = data["Proteins"]
            if proteins:
                ws.cell(row=row, column=1, value="Protein")
                ws.cell(row=row, column=2, value="Codon:Aminoacid")
                ws.cell(row=row, column=3, value="Stop Codon")
                row += 1

                for protein, pairs, stop_codon in proteins:
                    ws.cell(row=row, column=1, value=protein)
                    codon_str = " ".join(f"{c}:{aa}" for c, aa in pairs)
                    ws.cell(row=row, column=2, value=codon_str)
                    ws.cell(row=row, column=3, value=stop_codon or "—")
                    row += 2

            # Ajustare automat coloane
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = min(max_length + 2, 100)

        wb.save(filepath)
        return True
    except Exception as e:
        print(f"[export_xlsx] Failed: {e}")
        return False

# –––––––– GUI ––––––––
class ProteinTranslatorGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.selected_path = ""
        root.title("Protein Translator")
        root.geometry("500x320")
        root.resizable(False, False)
        self._build_ui()

    def _build_ui(self):
        pad = {"padx": 10, "pady": 4}
        tk.Label(self.root, text="Select input type:").pack(anchor="w", **pad)
        self.input_type = tk.StringVar(value="file")
        for text, val in (("Single File", "file"), ("Folder", "folder")):
            tk.Radiobutton(self.root, text=text, variable=self.input_type,
                           value=val).pack(anchor="w", padx=24)

        tk.Button(self.root, text="Browse", command=self._browse, width=12).pack(**pad)
        self.path_label = tk.Label(self.root, text="No file/folder selected",
                                   fg="blue", wraplength=460, justify="left")
        self.path_label.pack(**pad)

        tk.Label(self.root, text="Choose translation type:").pack(anchor="w", **pad)
        self.translation_type = tk.StringVar(value="both")
        for text, val in (("DNA → RNA + Proteins", "both"),
                          ("DNA → RNA only", "rna"),
                          ("Proteins only", "protein")):
            tk.Radiobutton(self.root, text=text, variable=self.translation_type,
                           value=val).pack(anchor="w", padx=24)

        tk.Button(self.root, text="Translate DNA", command=self._translate,
                  width=16).pack(pady=10)

    def _browse(self):
        if self.input_type.get() == "file":
            path = filedialog.askopenfilename(filetypes=[("Supported files", "*.txt *.csv *.fasta *.fa")])
        else:
            path = filedialog.askdirectory()
        if path:
            self.selected_path = path
            self.path_label.config(text=path)

    def _translate(self):
        if not self.selected_path:
            messagebox.showwarning("Warning", "Please select a file or folder first.")
            return

        translation_type = self.translation_type.get()
        if self.input_type.get() == "file":
            files = {os.path.basename(self.selected_path): self.selected_path}
        else:
            supported = [f for f in os.listdir(self.selected_path)
                         if f.lower().endswith((".txt", ".csv", ".fasta", ".fa"))]
            if not supported:
                messagebox.showerror("Error", "No supported files found in folder.")
                return
            files = {f: os.path.join(self.selected_path, f) for f in supported}

        results_dict = {}
        if len(files) == 1:
            name, path = next(iter(files.items()))
            results_dict[name] = process_file(path, translation_type)
        else:
            with ThreadPoolExecutor() as executor:
                futures = {name: executor.submit(process_file, path, translation_type)
                           for name, path in files.items()}
                results_dict = {name: fut.result() for name, fut in futures.items()}

        self._show_results(results_dict)

    def _show_results(self, results_dict: dict):
        win = tk.Toplevel(self.root)
        win.title("Translation Results")
        win.geometry("640x480")

        # Legend
        legend = tk.Frame(win)
        legend.pack(fill=tk.X, padx=8, pady=(6, 0))
        tk.Label(legend, text="Legend:").pack(side=tk.LEFT)
        tk.Label(legend, text=" AUG (start) ", bg="#2ecc71", fg="white",
                 font=("Courier", 9, "bold")).pack(side=tk.LEFT, padx=4)
        tk.Label(legend, text=" UAA/UAG/UGA (stop) ", bg="#e74c3c", fg="white",
                 font=("Courier", 9, "bold")).pack(side=tk.LEFT, padx=4)

        txt = scrolledtext.ScrolledText(win, wrap=tk.WORD, font=("Courier", 10))
        txt.pack(expand=True, fill=tk.BOTH, padx=6, pady=6)

        txt.tag_configure("start_codon", background="#2ecc71", foreground="white",
                          font=("Courier", 10, "bold"))
        txt.tag_configure("stop_codon",  background="#e74c3c", foreground="white",
                          font=("Courier", 10, "bold"))
        txt.tag_configure("header",      foreground="#2c3e50", font=("Courier", 10, "bold"))

        def insert_colored(text: str):
            tokens = text.split()
            for i, token in enumerate(tokens):
                space = "" if i == 0 else " "
                clean = token.strip("(),[]'\"")
                if clean == "AUG":
                    txt.insert(tk.END, space + token, "start_codon")
                elif clean in _STOP_CODONS:
                    txt.insert(tk.END, space + token, "stop_codon")
                else:
                    txt.insert(tk.END, space + token)
            txt.insert(tk.END, "\n")

        for filename, data in results_dict.items():
            txt.insert(tk.END, f"─── {filename} ───\n", "header")
            if data["RNA"]:
                txt.insert(tk.END, "RNA: ")
                insert_colored(data["RNA"])
            for idx, (protein, pairs, stop_codon) in enumerate(data["Proteins"], 1):
                txt.insert(tk.END, f"\nProtein {idx}: ")
                insert_colored(protein)
                txt.insert(tk.END, "Codons: ")
                insert_colored(" ".join(f"{c}:{aa}" for c, aa in pairs))
                if stop_codon:
                    txt.insert(tk.END, "Stop codon: ")
                    txt.insert(tk.END, stop_codon + "\n", "stop_codon")
            txt.insert(tk.END, "\n" + "═" * 50 + "\n\n")

        txt.configure(state="disabled")
        tk.Button(win, text="Export XLSX",
                  command=lambda: self._export_xlsx(results_dict)).pack(pady=6)

    def _export_xlsx(self, results_dict: dict):
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not filepath:
            return
        if export_xlsx(results_dict, filepath):
            messagebox.showinfo("Success", f"Exported to:\n{filepath}")
        else:
            messagebox.showerror("Error", "Export failed. Check console for details.")

# –––––––– RUN ––––––––
if __name__ == "__main__":
    root = tk.Tk()
    ProteinTranslatorGUI(root)
    root.mainloop()